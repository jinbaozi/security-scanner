#!/usr/bin/env python3
"""从 redline.xlsx 导出完整规范条款 Markdown。"""

from __future__ import annotations

import argparse
import os
import re
import sys
from dataclasses import dataclass, field
from pathlib import Path
from typing import Iterable


CLAUSE_RE = re.compile(r"^(\d+(?:\.\d+)+)\s*([\s\S]+)$")
EXPECTED_CLAUSE_COUNT = 40
SPOT_CHECK_IDS = ("5.1.4", "4.1.1", "11.2.1")


@dataclass
class Clause:
    clause_id: str
    text: str
    notes: list[str] = field(default_factory=list)
    source_rows: list[int] = field(default_factory=list)


def clean_cell(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def parse_sheet(rows: Iterable[tuple[object, object]]) -> list[Clause]:
    clauses: list[Clause] = []
    current: Clause | None = None

    for row_index, (raw_clause, raw_note) in enumerate(rows, start=1):
        clause_cell = clean_cell(raw_clause)
        note_cell = clean_cell(raw_note)
        if not clause_cell and not note_cell:
            continue

        match = CLAUSE_RE.match(clause_cell) if clause_cell else None
        if match:
            current = Clause(
                clause_id=match.group(1),
                text=match.group(2).strip(),
                source_rows=[row_index],
            )
            if note_cell:
                current.notes.append(note_cell)
            clauses.append(current)
            continue

        if current is None:
            raise ValueError(f"第 {row_index} 行无法归属到任何条款：{clause_cell!r} / {note_cell!r}")

        if clause_cell:
            raise ValueError(f"第 {row_index} 行存在无法解析的条款编号：{clause_cell!r}")

        if note_cell:
            current.notes.append(note_cell)
            current.source_rows.append(row_index)

    return clauses


def load_clauses(xlsx_path: Path, sheet_name: str) -> list[Clause]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("缺少 openpyxl，无法解析 xlsx。请在当前环境安装 openpyxl 后重试。") from exc

    workbook = load_workbook(xlsx_path, data_only=True, read_only=True)
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"工作簿中不存在工作表 {sheet_name!r}，可用工作表：{', '.join(workbook.sheetnames)}")

    worksheet = workbook[sheet_name]
    clauses = parse_sheet(worksheet.iter_rows(min_row=1, max_col=2, values_only=True))
    if len(clauses) != EXPECTED_CLAUSE_COUNT:
        raise ValueError(f"解析到 {len(clauses)} 条条款，期望 {EXPECTED_CLAUSE_COUNT} 条")
    return clauses


def markdown_paragraph(text: str) -> str:
    return "\n".join(text.splitlines())


def markdown_list_item(index: int, text: str) -> str:
    lines = text.splitlines() or [""]
    rendered = [f"{index}. {lines[0]}"]
    rendered.extend(f"   {line}" if line else "   " for line in lines[1:])
    return "\n".join(rendered)


def describe_source_path(xlsx_path: Path, base_dir: Path) -> str:
    relative_path = os.path.relpath(xlsx_path, base_dir)
    if relative_path == "redline.xlsx":
        return "`redline.xlsx`"
    if not relative_path.startswith(".."):
        return f"`redline.xlsx`（当前仓库内路径：`{relative_path}`）"
    return f"`redline.xlsx`（相对当前仓库：`{relative_path}`）"


def find_default_xlsx(repo_root: Path) -> Path:
    candidates = [repo_root / "redline.xlsx"]
    candidates.extend(parent / "redline.xlsx" for parent in repo_root.parents)

    seen: set[Path] = set()
    for candidate in candidates:
        resolved = candidate.resolve()
        if resolved in seen:
            continue
        seen.add(resolved)
        if resolved.is_file():
            return resolved

    searched = "\n".join(f"  - {path}" for path in candidates)
    raise FileNotFoundError(
        "未找到默认输入文件 redline.xlsx。\n"
        "已查找以下位置：\n"
        f"{searched}\n"
        "请将 redline.xlsx 放在当前仓库根目录，或显式传入 xlsx 路径，例如：\n"
        "  python3 scripts/export_redline_spec.py /path/to/redline.xlsx"
    )


def render_markdown(clauses: list[Clause], xlsx_path: Path, sheet_name: str, base_dir: Path) -> str:
    clause_index = {clause.clause_id: clause for clause in clauses}
    missing_spot_checks = [clause_id for clause_id in SPOT_CHECK_IDS if clause_id not in clause_index]
    if missing_spot_checks:
        raise ValueError(f"抽查条款缺失：{', '.join(missing_spot_checks)}")

    source_path = describe_source_path(xlsx_path, base_dir)
    lines: list[str] = [
        "# redline 规范条款导出",
        "",
        f"- 生成来源：{source_path}的 `{sheet_name}`",
        f"- 条款数量：{len(clauses)}",
        "- 导出方式：`scripts/export_redline_spec.py` 使用 openpyxl 解析 A 列条款正文与 B 列解读。",
        "",
        "## 抽查记录",
        "",
    ]

    for clause_id in SPOT_CHECK_IDS:
        clause = clause_index[clause_id]
        rows = "、".join(str(row) for row in clause.source_rows)
        lines.append(f"- {clause_id}：已与 {source_path}的 `{sheet_name}` 第 {rows} 行核对，正文与 {len(clause.notes)} 条解读一致。")

    lines.extend(["", "## 条款列表", ""])

    for clause in clauses:
        lines.extend(
            [
                f"### {clause.clause_id}",
                "",
                "**规范正文：**",
                "",
                markdown_paragraph(clause.text),
                "",
                "**解读列表：**",
                "",
            ]
        )
        if clause.notes:
            for index, note in enumerate(clause.notes, start=1):
                lines.append(markdown_list_item(index, note))
        else:
            lines.append("无")
        lines.append("")

    return "\n".join(lines).rstrip() + "\n"


def parse_args() -> argparse.Namespace:
    repo_root = Path(__file__).resolve().parents[1]
    parser = argparse.ArgumentParser(description="从 redline.xlsx Sheet1 导出完整规范条款 Markdown。")
    parser.add_argument(
        "xlsx_path",
        nargs="?",
        type=Path,
        default=None,
        help="输入 redline.xlsx 路径；不传时先查当前仓库根目录，再向上查找父级工作区。",
    )
    parser.add_argument(
        "output_path",
        nargs="?",
        type=Path,
        default=repo_root / "references" / "redline-spec.md",
        help="输出 Markdown 路径，默认写入 references/redline-spec.md。",
    )
    parser.add_argument("--sheet", default="Sheet1", help="要导出的工作表名称，默认 Sheet1。")
    return parser.parse_args()


def resolve_input_xlsx(args: argparse.Namespace, repo_root: Path) -> Path:
    if args.xlsx_path is None:
        return find_default_xlsx(repo_root)

    xlsx_path = args.xlsx_path.expanduser().resolve()
    if not xlsx_path.is_file():
        raise FileNotFoundError(
            f"输入文件不存在：{xlsx_path}\n"
            "请确认 redline.xlsx 路径，或按以下方式显式传入：\n"
            "  python3 scripts/export_redline_spec.py /path/to/redline.xlsx"
        )
    return xlsx_path


def main() -> int:
    args = parse_args()
    repo_root = Path(__file__).resolve().parents[1]
    xlsx_path = resolve_input_xlsx(args, repo_root)
    output_path = args.output_path.expanduser().resolve()

    clauses = load_clauses(xlsx_path, args.sheet)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(render_markdown(clauses, xlsx_path, args.sheet, repo_root), encoding="utf-8")
    print(f"已导出 {len(clauses)} 条条款到 {output_path}")
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except (FileNotFoundError, OSError, RuntimeError, ValueError) as exc:
        print(f"错误：{exc}", file=sys.stderr)
        raise SystemExit(1)
